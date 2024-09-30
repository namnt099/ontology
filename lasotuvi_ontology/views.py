import datetime
import json
from re import template
from unicodedata import name
from cv2 import CAP_OPENNI2
from numpy import ndenumerate
import pyodbc 
import pandas as pd
from colorama import init, Fore, Back, Style
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment
from openpyxl import Workbook, load_workbook, worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from django.http import HttpResponse
from django.shortcuts import render
from lasotuvi.DiaBan import cungDiaBan, diaBan, lapDiaBan
from lasotuvi.ThienBan import lapThienBan
from lasotuvi.AmDuong import ngayThangNamCanChi
# from lasotuvi.AmDuong import ngayThangNam
import openpyxl
from copy import copy
from django.views.generic import View
from .process import html_to_pdf 
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import StringIO, BytesIO
import urllib.request, urllib.error, urllib.parse
import numpy as np
import json
import re
import pandas as pd
from owlready2 import *
from owlready2 import get_ontology, default_world

db = {'diaban': '', 'thienban': ''}

def lasotuvi_django_index(request):
    return render(request, 'index.html')

class GeneratePdf(View):
    def get(self, request, *args, **kwargs):
        pdf = html_to_pdf('index.html')
        return HttpResponse(pdf, content_type='application/pdf')




# Sua lai loi ko nhan lich am
def api_laso(request):
    
    now = datetime.datetime.now()
    hoTen = (request.GET.get('hoten'))
    ngaySinh = int(request.GET.get('ngaysinh', now.day))
    thangSinh = int(request.GET.get('thangsinh', now.month))
    namSinh = int(request.GET.get('namsinh', now.year))
    gioiTinh = 1 if request.GET.get('gioitinh') == 'nam' else -1
    gioSinh = int(request.GET.get('giosinh', 1))
    timeZone = int(request.GET.get('muigio', 7))
    namXemHan = int(request.GET.get('namxemhan', now.year))
    # thamkhao = list(request.GET.getlist('thamkhao'))
    thamkhao = request.GET.get('tailieuthamkhao')
    duongLich = False if request.GET.get('amlich') == 'on' else True
    if db["diaban"] != "":
        db["diaban"]= ""
        db["thienban"] = ""
        # db["ntmcc"] = ""
    db["diaban"] = lapDiaBan(diaBan, ngaySinh, thangSinh, namSinh, gioSinh,
                   gioiTinh, namXemHan, duongLich, timeZone)
    db["thienban"]  = lapThienBan(ngaySinh, thangSinh, namSinh,
                           gioSinh, gioiTinh, hoTen, db["diaban"], namXemHan, duongLich)
    ntmcc = ngayThangNamCanChi(ngaySinh, thangSinh, namSinh, timeZone)
    # Mapping of zodiac signs to their order and vice versa for easy lookups.
    zodiac_signs = ['Tí', 'Sửu', 'Dần', 'Mão', 'Thìn', 'Tỵ', 'Ngọ', 'Mùi', 'Thân', 'Dậu', 'Tuất', 'Hợi']
    sign_to_index = {sign: index for index, sign in enumerate(zodiac_signs)}

    # Starting sign based on the remainder of ntmcc divided by 4
    start_signs = ['Sửu', 'Tuất', 'Mùi', 'Thìn']
    start_sign = start_signs[ntmcc[2] % 4]
    
    def generate_tieuHan(df, start_sign, gioiTinh):
        start_index = sign_to_index[start_sign]
        num_signs = len(zodiac_signs)
        tieuHan_values = []
        for sign in df['CungSo']:
            current_index = sign_to_index[sign]
            if gioiTinh == 1:
                distance = (current_index - start_index) % num_signs
            else:
                distance = (start_index - current_index) % num_signs
            tieuHan_values.append(f'T{distance + 1}')
        return tieuHan_values
    
    cungSo_mapping = {
        1: 'Tí',
        2: 'Sửu',
        3: 'Dần',
        4: 'Mão',
        5: 'Thìn',
        6: 'Tỵ',
        7: 'Ngọ',
        8: 'Mùi',
        9: 'Thân',
        10: 'Dậu',
        11: 'Tuất',
        12: 'Hợi'
    }
    
    dacTinhSao_mapping = {
        'H': 'Hãm',
        'Đ': 'Đắc',
        'M': 'Miếu',
        'V': 'Vượng',
        'B': 'Bình'
    }
    
    diaban = db["diaban"]
    thienban = db["thienban"]
    global a, b, tk
    a = db["diaban"].thapNhiCung
    b = db["thienban"]

    tk = thamkhao

    diaban_data = []
    vanhan_data = []
    daihan_data = []
    for cung in diaban.thapNhiCung:
        cungSo = cungSo_mapping.get(cung.cungSo, cung.cungSo)
        cungChu = getattr(cung, 'cungChu', '')
        cungChu = cungChu.title() if isinstance(cungChu, str) else cungChu  
        cungDaiHan = getattr(cung, 'cungDaiHan', '')
        # print(cungDaiHan)
        
        for sao in cung.cungSao:
            daihan_data.append({
                    'CungSo': cungSo,
                    'CungChu': cungChu,
                    'DaiHan': cungDaiHan,
                })
            saoTen = sao['saoTen'].title()
            if 'L.' in saoTen:
                # Sao có tiền tố "L.", lưu vào danh sách vanhan_data
                dac_tinh = sao['saoDacTinh'] if 'saoDacTinh' in sao else ''
                mapped_dac_tinh = dacTinhSao_mapping.get(dac_tinh, dac_tinh)  

                vanhan_data.append({
                    'CungSo': cungSo,
                    'CungChu': cungChu,
                    'SaoTen': saoTen,
                })
            else:
                # Sao không có tiền tố "L.", lưu vào danh sách diaban_data
                dac_tinh = sao['saoDacTinh'] if 'saoDacTinh' in sao else ''
                mapped_dac_tinh = dacTinhSao_mapping.get(dac_tinh, dac_tinh)  

                diaban_data.append({
                    'CungSo': cungSo,
                    'CungChu': cungChu,
                    'SaoTen': saoTen,
                    'dacTinhSao': mapped_dac_tinh
                })

    # Lưu dữ liệu vào 'diaban.csv'
    df_diaban = pd.DataFrame(diaban_data)
    df_diaban.to_csv('diaban.csv', header=None, index=False, encoding='utf-8-sig')

    df_daihan = pd.DataFrame(daihan_data)
    df_daihan.drop_duplicates(inplace=True)
    df_daihan.to_csv('daihan.csv', header=None, index=False, encoding='utf-8-sig')

    # Lưu dữ liệu vào 'vanhan.csv' cho các sao có tiền tố "L."
    df_vanhan = pd.DataFrame(vanhan_data)
    df_vanhan['tieuHan'] = generate_tieuHan(df_vanhan, start_sign, gioiTinh)
    # print(df_vanhan)
    df_vanhan.to_csv('vanhan.csv', header=None, index=False, encoding='utf-8-sig')

    # Tạo và lưu dữ liệu cho 'thienban.csv'
    am_duong_menh = ' '.join(thienban.amDuongMenh.split()[-2:]).capitalize()
    thienban_data = {
        # 'Triệt': ['Triệt'] + [cung.cungTen for cung in diaban.thapNhiCung if cung.trietLo],
        # 'Tuần': ['Tuần'] + [cung.cungTen for cung in diaban.thapNhiCung if cung.tuanTrung],
        # 'Can': ['Can', 'Người', thienban.canNamTen],
        # 'Chi': ['Chi', 'Người', thienban.chiNamTen],
        # 'Âm Dương': ['Âm Dương', 'Người', am_duong_menh],
        # 'Thân': ['Cung Thân'] + [cung.cungChu for cung in diaban.thapNhiCung if cung.cungThan]
        'Triệt': ['Triệt'] + [cung.cungTen for cung in diaban.thapNhiCung if getattr(cung, 'trietLo', False)],
        'Tuần': ['Tuần'] + [cung.cungTen for cung in diaban.thapNhiCung if getattr(cung, 'tuanTrung', False)],
        'Can': ['Can', 'Người', thienban.canNamTen],
        'Chi': ['Chi', 'Người', thienban.chiNamTen],
        'Âm Dương': ['Âm Dương', 'Người', am_duong_menh],
        'Thân': ['Cung Thân'] + [cung.cungChu for cung in diaban.thapNhiCung if getattr(cung, 'cungThan', False)]
    }

    # Lưu dữ liệu vào 'thienban.csv'
    df_thienban = pd.DataFrame.from_dict(thienban_data, orient='index')
    df_thienban.to_csv('thienban.csv', header=None, index=False, encoding='utf-8-sig')

    laso = {
        'thienBan': db["thienban"],
        'thapNhiCung': db["diaban"].thapNhiCung,
    }
    my_return = (json.dumps(laso, default=lambda o: o.__dict__))
    return HttpResponse(my_return, content_type="application/json")

def unique_list(l_2d):
    kiemtra = []
    return [x for x in l_2d if x not in kiemtra and not kiemtra.append(x)]
def process_ontology_data(df_dia_ban, df_thien_ban, df_im):
    onto = get_ontology("id.rdf")
    with onto:
        class Individuals(Thing):
            pass
        class DongCung(ObjectProperty):
            domain    = [Individuals]
            range     = [Individuals]
        class CungDongCung(ObjectProperty):
            domain    = [Individuals]
            range     = [Individuals]
        class Chua(ObjectProperty):
            domain    = [Individuals]
            range     = [Individuals]
        class DacTinh(ObjectProperty):
            domain    = [Individuals]
            range     = [Individuals]
        class Giap(ObjectProperty,SymmetricProperty):
            domain    = [Individuals]
            range     = [Individuals]
        class TamHop(ObjectProperty,SymmetricProperty):
            domain    = [Individuals]
            range     = [Individuals]
        class NhiHop(ObjectProperty,SymmetricProperty):
            domain    = [Individuals]
            range     = [Individuals]
        class XungChieu(ObjectProperty,SymmetricProperty):
            domain    = [Individuals]
            range     = [Individuals]
        class Tuan(ObjectProperty,SymmetricProperty):
            domain    = [Individuals]
            range     = [Individuals]
        class Triet(ObjectProperty,SymmetricProperty):
            domain    = [Individuals]
            range     = [Individuals]
        class Can(ObjectProperty):
            domain    = [Individuals]
            range     = [Individuals]
        class Chi(ObjectProperty):
            domain    = [Individuals]
            range     = [Individuals]
        class AmDuong(ObjectProperty):
            domain    = [Individuals]
            range     = [Individuals]



    for j in range(0,len(df_dia_ban)):
        label1 = str(df_dia_ban.loc[j,1])
        label2 = str(df_dia_ban.loc[j,0])
        label3 = str(df_dia_ban.loc[j,2])
        label4 = str(df_dia_ban.loc[j,3])
        name1 = label1.replace(" ","")
        name2 = label2.replace(" ","")
        name3 = label3.replace(" ","")
        name4 = label4.replace(" ","")
        indi1 = Individuals(name1)
        if len(list(indi1.label))==0:
            indi1.label.append(label1)
        indi2 = Individuals(name2)
        if len(list(indi2.label))==0:
            indi2.label.append(label2)
        indi3 = Individuals(name3)
        if len(list(indi3.label))==0:
            indi3.label.append(label3)

        if indi2 not in indi1.DongCung:
            indi1.DongCung.append(indi2)
        indi1.Chua.append(indi3)

        if name4!="nan":
            indi4 = Individuals(name4)
            if len(list(indi4.label))==0:
                indi4.label.append(label4)
                indi3.DacTinh.append(indi4)

    tamhop = []
    xungchieu = []
    cung = []

    for j in range(0,len(df_dia_ban)):
        if str(df_dia_ban.loc[j,1]) not in cung:
            cung.append(str(df_dia_ban.loc[j,1]))

    # giap
    for j in range(0,len(cung)-1,2):
        if j==0:
            label1 = cung[j]
            label2 = cung[len(cung)-1]
            label3 = cung[j+1]
        else:
            label1 = cung[j]
            label2 = cung[j-1]
            label3 = cung[j+1]
        name1 = label1.replace(" ","")
        name2 = label2.replace(" ","")
        name3 = label3.replace(" ","")
        indi1 = Individuals(name1)
        indi2 = Individuals(name2)
        indi3 = Individuals(name3)
        indi1.Giap.append(indi2)
        indi1.Giap.append(indi3)

    # tam hợp
    for j in range(0,len(cung)):
        if cung[j] not in tamhop:
            if j+8<=len(cung):
                label1 = cung[j]
                label2 = cung[j+4]
                label3 = cung[j+8]
                tamhop.append(label1)
                tamhop.append(label2)
                tamhop.append(label3)
                name1 = label1.replace(" ","")
                name2 = label2.replace(" ","")
                name3 = label3.replace(" ","")
                indi1 = Individuals(name1)
                indi2 = Individuals(name2)
                indi3 = Individuals(name3)
                indi1.TamHop.append(indi2)
                indi1.TamHop.append(indi3)
                indi2.TamHop.append(indi3)

    # nhi hop
    for j in range(6):
        label1 = cung[j]
        label2 = cung[len(cung)-1-j]
        name1 = label1.replace(" ","")
        name2 = label2.replace(" ","")
        indi1 = Individuals(name1)
        indi2 = Individuals(name2)
        indi1.NhiHop.append(indi2)

        temp = cung[0]
        cung.remove(cung[0])
        cung.append(temp)

    # xung chieu
    for j in range(0,len(cung)):
        if cung[j] not in xungchieu:
            if j+6<=len(cung):
                label1 = cung[j]
                label2 = cung[j+6]
                xungchieu.append(label1)
                xungchieu.append(label2)
                name1 = label1.replace(" ","")
                name2 = label2.replace(" ","")
                indi1 = Individuals(name1)
                indi2 = Individuals(name2)
                indi1.XungChieu.append(indi2)

    name1 = str(df_thien_ban.loc[0,1]).replace(" ","")
    name2 = str(df_thien_ban.loc[0,2]).replace(" ","")
    indi1 = Individuals(name1)
    indi2 = Individuals(name2)
    indi1.Triet.append(indi2)
    if len(list(indi1.label))==0:
        indi1.label.append(str(df_thien_ban.loc[0,1]))
    if len(list(indi2.label))==0:
        indi2.label.append(str(df_thien_ban.loc[0,2]))

    name1 = str(df_thien_ban.loc[1,1]).replace(" ","")
    name2 = str(df_thien_ban.loc[1,2]).replace(" ","")
    indi1 = Individuals(name1)
    indi2 = Individuals(name2)
    indi1.Tuan.append(indi2)
    if len(list(indi1.label))==0:
        indi1.label.append(str(df_thien_ban.loc[1,1]))
    if len(list(indi2.label))==0:
        indi2.label.append(str(df_thien_ban.loc[1,2]))

    name1 = str(df_thien_ban.loc[2,1]).replace(" ","")
    name2 = str(df_thien_ban.loc[2,2]).replace(" ","")
    indi1 = Individuals(name1)
    indi2 = Individuals(name2)
    indi1.Can.append(indi2)
    if len(list(indi1.label))==0:
        indi1.label.append(str(df_thien_ban.loc[2,1]))
    if len(list(indi2.label))==0:
        indi2.label.append(str(df_thien_ban.loc[2,2]))

    name1 = str(df_thien_ban.loc[3,1]).replace(" ","")
    name2 = str(df_thien_ban.loc[3,2]).replace(" ","")
    indi1 = Individuals(name1)
    indi2 = Individuals(name2)
    indi1.Chi.append(indi2)
    if len(list(indi1.label))==0:
        indi1.label.append(str(df_thien_ban.loc[3,1]))
    if len(list(indi2.label))==0:
        indi2.label.append(str(df_thien_ban.loc[3,2]))

    name1 = str(df_thien_ban.loc[4,1]).replace(" ","")
    name2 = str(df_thien_ban.loc[4,2]).replace(" ","")
    indi1 = Individuals(name1)
    indi2 = Individuals(name2)
    indi1.AmDuong.append(indi2)
    if len(list(indi1.label))==0:
        indi1.label.append(str(df_thien_ban.loc[4,1]))
    if len(list(indi2.label))==0:
        indi2.label.append(str(df_thien_ban.loc[4,2]))
    
    name1 = str(df_thien_ban.loc[5,0]).replace(" ","")
    name2 = str(df_thien_ban.loc[5,1]).replace(" ","")
    indi1 = Individuals(name1)
    indi2 = Individuals(name2)
    indi1.CungDongCung.append(indi2)
    if len(list(indi1.label))==0:
        indi1.label.append(str(df_thien_ban.loc[5,0]))
    if len(list(indi2.label))==0:
        indi2.label.append(str(df_thien_ban.loc[5,1]))

    sync_reasoner()


    """## Query 1: Đóng tại cung nào?"""
    def DongCung(search_term):
        pre_query = """?x rdf:type owl:NamedIndividual. ?x <id.rdf#DongCung> ?y. ?y rdfs:label ?ylabel. ?x rdfs:label "{keyword}".""".format(keyword=str(search_term))
        query = """SELECT DISTINCT (STR(?ylabel) AS ?label)"""+ "{" + pre_query + "}"
        query_result = list(default_world.sparql(str(query)))
        result = []
        for i in range(len(query_result)):
            for j in range(len(query_result[i])):
                result.append(query_result[i][j])
        return result
    
    def CungDongCung():
        pre_query = """?x rdf:type owl:NamedIndividual. ?x <id.rdf#CungDongCung> ?y. ?y rdfs:label ?ylabel. ?x rdfs:label "{keyword}".""".format(keyword="Cung Thân")
        query = """SELECT DISTINCT (STR(?ylabel) AS ?label)"""+ "{" + pre_query + "}"
        query_result = list(default_world.sparql(str(query)))
        result = []
        for i in range(len(query_result)):
            for j in range(len(query_result[i])):
                result.append(query_result[i][j])
        return result


    """## Query 2: Một cung chứa những sao nào?"""
    def Chua(search_term):
        pre_query = """?x rdf:type owl:NamedIndividual. ?x <id.rdf#Chua> ?y. ?y rdfs:label ?ylabel. ?x rdfs:label "{keyword}".""".format(keyword=str(search_term))
        query = """SELECT DISTINCT (STR(?ylabel) AS ?label)"""+ "{" + pre_query + "}"
        query_result = list(default_world.sparql(str(query)))
        result = []
        for i in range(len(query_result)):
            for j in range(len(query_result[i])):
                result.append(query_result[i][j])
        return result

    def NamTrongCung(search_term):
        pre_query = """?x rdf:type owl:NamedIndividual. ?y <id.rdf#Chua> ?x. ?y rdfs:label ?ylabel. ?x rdfs:label "{keyword}".""".format(keyword=str(search_term))
        query = """SELECT DISTINCT (STR(?ylabel) AS ?label)"""+ "{" + pre_query + "}"
        query_result = list(default_world.sparql(str(query)))
        result = []
        for i in range(len(query_result)):
            for j in range(len(query_result[i])):
                result.append(query_result[i][j])
        return result

    def TaiCung(search_term):
        pre_query = """?x rdf:type owl:NamedIndividual. ?z <id.rdf#Chua> ?x. ?z <id.rdf#DongCung> ?y. ?y rdfs:label ?ylabel. ?x rdfs:label "{keyword}".""".format(keyword=str(search_term))
        query = """SELECT DISTINCT (STR(?ylabel) AS ?label)"""+ "{" + pre_query + "}"
        query_result = list(default_world.sparql(str(query)))
        result = []
        for i in range(len(query_result)):
            for j in range(len(query_result[i])):
                result.append(query_result[i][j])
        return result

    def CungChua(search_term):
        pre_query = """?x rdf:type owl:NamedIndividual. ?z <id.rdf#Chua> ?y. ?z <id.rdf#DongCung> ?x. ?y rdfs:label ?ylabel. ?x rdfs:label "{keyword}".""".format(keyword=str(search_term))
        query = """SELECT DISTINCT (STR(?ylabel) AS ?label)"""+ "{" + pre_query + "}"
        query_result = list(default_world.sparql(str(query)))
        result = []
        for i in range(len(query_result)):
            for j in range(len(query_result[i])):
                result.append(query_result[i][j])
        return result

    """## Query 3: Đặc tính của Sao?"""
    def DacTinh(search_term):
        pre_query = """?x rdf:type owl:NamedIndividual. ?x <id.rdf#DacTinh> ?y. ?y rdfs:label ?ylabel. ?x rdfs:label "{keyword}".""".format(keyword=str(search_term))
        query = """SELECT DISTINCT (STR(?ylabel) AS ?label)"""+ "{" + pre_query + "}"
        query_result = list(default_world.sparql(str(query)))
        result = []
        for i in range(len(query_result)):
            for j in range(len(query_result[i])):
                result.append(query_result[i][j])
        return result

    """## Query 4: Tam hợp"""
    def TamHop(search_term):
        pre_query = """?x rdf:type owl:NamedIndividual. {{?x <id.rdf#TamHop> ?y}} UNION {{?y <id.rdf#TamHop> ?x}}.?x rdfs:label "{keyword}". ?y rdfs:label ?ylabel.""".format(keyword=str(search_term))
        query = """SELECT DISTINCT (STR(?ylabel) AS ?label)"""+ "{" + pre_query + "}"
        query_result = list(default_world.sparql(str(query)))
        result = []
        for i in range(len(query_result)):
            for j in range(len(query_result[i])):
                result.append(query_result[i][j])
        return result

    """##Query 5: Nhị hợp"""
    def NhiHop(search_term):
        pre_query = """?x rdf:type owl:NamedIndividual. {{?x <id.rdf#NhiHop> ?y}} UNION {{?y <id.rdf#NhiHop> ?x}}. ?x rdfs:label "{keyword}". ?y rdfs:label ?ylabel.""".format(keyword=str(search_term))
        query = """SELECT DISTINCT (STR(?ylabel) AS ?label)"""+ "{" + pre_query + "}"
        query_result = list(default_world.sparql(str(query)))
        result = []
        for i in range(len(query_result)):
            for j in range(len(query_result[i])):
                result.append(query_result[i][j])
        return result

    """##Query 6: Xung chiếu"""
    def XungChieu(search_term):
        pre_query = """?x rdf:type owl:NamedIndividual. {{?x <id.rdf#XungChieu> ?y}} UNION {{?y <id.rdf#XungChieu> ?x}}. ?x rdfs:label "{keyword}". ?y rdfs:label ?ylabel.""".format(keyword=str(search_term))
        query = """SELECT DISTINCT (STR(?ylabel) AS ?label)"""+ "{" + pre_query + "}"
        query_result = list(default_world.sparql(str(query)))
        result = []
        for i in range(len(query_result)):
            for j in range(len(query_result[i])):
                result.append(query_result[i][j])
        return result

    """##Query 7: Can, Chi, Âm Dương"""
    def CanChi():
        search_term = "Người"
        pre_query = """?x rdf:type owl:NamedIndividual. ?x <id.rdf#Can> ?y. ?x <id.rdf#Chi> ?z. ?x <id.rdf#AmDuong> ?t. ?x rdfs:label "{keyword}". ?y rdfs:label ?ylabel. ?z rdfs:label ?zlabel. ?t rdfs:label ?tlabel.""".format(keyword=str(search_term))
        query = """SELECT DISTINCT ?ylabel ?zlabel ?tlabel"""+ "{" + pre_query + "}"
        query_result = list(default_world.sparql(str(query)))
        result = []
        for i in range(len(query_result)):
            for j in range(len(query_result[i])):
                result.append(query_result[i][j])
        return result

    """## Query 8: Chiếu, Hội họp"""
    def Chieu(search_term):
        result = []
        result.append(search_term)
        result.extend(TamHop(search_term))
        result.extend(XungChieu(search_term))
        final_result = []
        for i in range(len(result)):
            final_result.extend(Chua(result[i]))
        return final_result

    def IsHoiHop(search_list):
        cung = NamTrongCung(search_list[0])
        temp = TamHop(cung[0])
        result1 = Chieu(cung[0])
        result2 = Chieu(temp[0])
        result3 = Chieu(temp[1])
        result_1 = set(search_list).issubset(result1)
        result_2 = set(search_list).issubset(result2)
        result_3 = set(search_list).issubset(result3)
        return result_1 or result_2 or result_3

    """## Query giáp"""
    def Giap(search_term):
        pre_query = """?x rdf:type owl:NamedIndividual. {{?x <id.rdf#Giap> ?y}} UNION {{?y <id.rdf#Giap> ?x}}.?x rdfs:label "{keyword}". ?y rdfs:label ?ylabel.""".format(keyword=str(search_term))
        query = """SELECT DISTINCT (STR(?ylabel) AS ?label)"""+ "{" + pre_query + "}"
        query_result = list(default_world.sparql(str(query)))
        result = []
        for i in range(len(query_result)):
            for j in range(len(query_result[i])):
                result.append(query_result[i][j])
        return result

    """## Query tuần, triệt"""

    def Tuan():
        pre_query = """{{?x <id.rdf#Tuan> ?y}} UNION {{?y <id.rdf#Tuan> ?x}}. ?x rdfs:label ?xlabel."""
        query = """SELECT DISTINCT ?xlabel"""+ "{" + pre_query + "}"
        query_result = list(default_world.sparql(str(query)))
        result = []
        for i in range(len(query_result)):
            for j in range(len(query_result[i])):
                result.append(query_result[i][j])
        return result

    def Triet():
        pre_query = """{{?x <id.rdf#Triet> ?y}} UNION {{?y <id.rdf#Triet> ?x}}. ?x rdfs:label ?xlabel."""
        query = """SELECT DISTINCT ?xlabel"""+ "{" + pre_query + "}"
        query_result = list(default_world.sparql(str(query)))
        result = []
        for i in range(len(query_result)):
            for j in range(len(query_result[i])):
                result.append(query_result[i][j])
        return result

    """## Kiểm tra đồng cung"""
    def IsDongCung(DanhSachSao):
        cung = NamTrongCung(DanhSachSao[0])
        DanhSachTrongCung = Chua(cung[0])
        if set(DanhSachSao).issubset(DanhSachTrongCung):
            return True
        return False

    """## Giáp với những sao nào"""

    def GiapSao(TenSao):
        DanhSachCung = Giap(NamTrongCung(TenSao)[0])
        result = []
        for i in range(len(DanhSachCung)):
            result.extend(Chua(DanhSachCung[i]))
        return result

    """## Information Extraction"""


    regex0 = [r'Có (.*) tại (.*), người tuổi (.*)']
    regex1 = [r'(.*) tại (.*) có (.*) gặp (.*) hội họp, người tuổi (.*)']
    regex2 = [r'(.*) tại (.*) có (.*) không gặp (.*), người tuổi (.*)']
    regex3 = [r'(.*) tại (.*) có (.*) gặp (.*), người tuổi (.*)']
    regex4 = [r'(.*) tại (.*) có (.*) chiếu, người tuổi (.*)']
    regex5 = [r'(.*) tại (.*) có (.*), người tuổi (.*)']
    regex6 = [r'(.*) có (.*) tại (.*) gặp (.*), người tuổi (.*)']
    regex7 = [r'(.*) có (.*) tại (.*), người tuổi (.*)']
    regex8 = [r'(.*) có (.*) gặp (.*) hội họp, người tuổi (.*)']
    regex9 = [r'(.*) có (.*) gặp (.*), người tuổi (.*)']
    regex10 = [r'(.*) có (.*), người tuổi (.*)']
    regex11 = [r'(.*) tại (.*) gặp (.*) và không gặp (.*), người tuổi (.*)']
    regex12 = [r'(.*) tại (.*) không gặp (.*) hội họp, người tuổi (.*)']
    regex13 = [r'(.*) tại (.*) không gặp (.*), người tuổi (.*)']
    regex14 = [r'(.*) tại (.*) gặp (.*) hội họp, người tuổi (.*)']
    regex15 = [r'(.*) tại (.*) gặp (.*), người tuổi (.*)']
    regex16 = [r'(.*) hội họp, người tuổi (.*)']
    regex17 = [r'Có (.*) tại (.*)']
    regex18 = [r'(.*) ở vượng địa']
    regex19 = [r'(.*) có (.*) đồng cung gặp (.*) chiếu']
    regex20 = [r'(.*) có (.*) đồng cung gặp (.*)']
    regex21 = [r'(.*) có (.*) gặp (.*) đồng cung chiếu']
    regex22 = [r'(.*) có (.*) không gặp (.*)']
    regex23 = [r'(.*) có (.*) tại (.*) giáp (.*)']
    regex24 = [r'(.*) có (.*) tại (.*) gặp (.*)']
    regex25 = [r'(.*) có (.*) tại (.*)']
    regex26 = [r'(.*) có (.*) gặp (.*) hội họp']
    regex27 = [r'(.*) có (.*) gặp (.*) chiếu']
    regex28 = [r'(.*) có (.*) gặp (.*)']
    regex29 = [r'(.*) có (.*) đồng cung']
    regex30 = [r'(.*) có (.*) chiếu']
    regex31 = [r'(.*) có (.*)']

    regex32 = [r'(.*) tại (.*) có (.*) không gặp (.*)']
    regex33 = [r'(.*) tại cung (.*) có (.*) gặp (.*)']
    regex34 = [r'(.*) tại (.*) có (.*) gặp (.*) hội họp']
    regex35 = [r'(.*) tại (.*) có (.*) gặp (.*) chiếu']
    regex36 = [r'(.*) tại (.*) có (.*) gặp (.*)']
    regex37 = [r'(.*) tại cung (.*) có (.*)']
    regex38 = [r'(.*) tại (.*) có (.*) chiếu']
    regex39 = [r'(.*) tại (.*) có (.*)']

    regex40 = [r'(.*) tại (.*) gặp (.*) và không gặp (.*)']
    regex41 = [r'(.*) gặp (.*) và không gặp (.*)']
    regex42 = [r'(.*) không gặp (.*)']
    regex43 = [r'(.*) tại cung (.*) gặp (.*) hội họp']
    regex44 = [r'(.*) tại cung (.*) gặp (.*)']
    regex45 = [r'(.*) tại cung (.*) giáp (.*)']
    regex46 = [r'(.*) tại (.*) gặp (.*) hội họp']
    regex47 = [r'(.*) tại (.*) gặp (.*) chiếu']
    regex48 = [r'(.*) tại (.*) gặp (.*)']
    regex49 = [r'(.*) tại (.*)']
    regex50 = [r'(.*) gặp (.*) hội họp']
    regex51 = [r'(.*) gặp (.*) chiếu']
    regex52 = [r'(.*) gặp (.*)']
    regex53 = [r'(.*) hội họp']
    regex54 = [r'(.*) đồng cung']
    
    regex55 = [r'(.*) có (.*), (.*) có (.*) gặp (.*)']
    regex56 = [r'(.*) có (.*), (.*) có (.*) gặp (.*) và (.*)']
    regex57 = [r'(.*) có (.*), (.*) có (.*) nếu gặp (.*)']
    regex58 = [r'(.*) có (.*), (.*) có (.*) nếu được (.*) hội họp']

    regex_list1 = regex0 + regex1 + regex2 + regex3 + regex4 + regex5 + regex6 + regex7 + regex8 + regex9 + regex10 + regex11 + regex12 + regex13 + regex14 + regex15 + regex16
    regex_list2 = regex17 + regex18
    regex_list4 = regex32 + regex33 + regex34 + regex35 + regex36 + regex37 + regex38 + regex39
    regex_list3 = regex19 + regex20 + regex21 + regex22 + regex23 + regex24 + regex25 + regex26 + regex27 + regex28 + regex29 + regex30 + regex31
    regex_list5 = regex40 + regex41 + regex42 + regex43 + regex44 + regex45 + regex46 + regex47 + regex48 + regex49 + regex50 + regex51 + regex52 + regex53 + regex54
    regex_list6 = regex55 + regex56 + regex57 + regex58
    regex_list = regex_list1 + regex_list2 + regex_list4 + regex_list3 + regex_list5 + regex_list6
    #len(regex_list)

    def extract(sentence):
        Cung_list = []
        Chi_list = []
        Sao_list1 = []
        Sao_list2 = []
        Sao_list3 = []
        Can_list = []
        for i in range(1):
            if re.match(regex_list[0], sentence):
                a = re.match(regex_list[0], sentence)
                Cung_list.append('')
                Chi_list.append(a[2])
                Sao_list1.append(a[1])
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append(a[3])
                continue
            elif re.match(regex_list[1], sentence):
                a = re.match(regex_list[1], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[2])
                Sao_list1.append(a[3])
                Sao_list2.append(a[4])
                Sao_list3.append('')
                Can_list.append(a[5])
                continue
            elif re.match(regex_list[2], sentence):
                a = re.match(regex_list[2], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[2])
                Sao_list1.append(a[3])
                Sao_list2.append('')
                Sao_list3.append(a[4])
                Can_list.append(a[5])
                continue
            elif re.match(regex_list[3], sentence):
                a = re.match(regex_list[3], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[2])
                Sao_list1.append(a[3])
                Sao_list2.append(a[4])
                Sao_list3.append('')
                Can_list.append(a[5])
                continue
            elif re.match(regex_list[4], sentence):
                a = re.match(regex_list[4], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[2])
                Sao_list1.append(a[3])
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append(a[4])
                continue
            elif re.match(regex_list[5], sentence):
                a = re.match(regex_list[5], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[2])
                Sao_list1.append(a[3])
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append(a[4])
                continue
            elif re.match(regex_list[6], sentence):
                a = re.match(regex_list[6], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[3])
                Sao_list1.append(a[2])
                Sao_list2.append(a[4])
                Sao_list3.append('')
                Can_list.append(a[5])
                continue
            elif re.match(regex_list[7], sentence):
                a = re.match(regex_list[7], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[3])
                Sao_list1.append(a[2])
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append(a[4])
                continue
            elif re.match(regex_list[8], sentence):
                a = re.match(regex_list[8], sentence)
                Cung_list.append(a[1])
                Chi_list.append('')
                Sao_list1.append(a[2])
                Sao_list2.append(a[3])
                Sao_list3.append('')
                Can_list.append(a[4])
                continue
            elif re.match(regex_list[9], sentence):
                a = re.match(regex_list[9], sentence)
                Cung_list.append(a[1])
                Chi_list.append('')
                Sao_list1.append(a[2])
                Sao_list2.append(a[3])
                Sao_list3.append('')
                Can_list.append(a[4])
                continue
            elif re.match(regex_list[10], sentence):
                a = re.match(regex_list[10], sentence)
                Cung_list.append(a[1])
                Chi_list.append('')
                Sao_list1.append(a[2])
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append(a[3])
                continue
            elif re.match(regex_list[11], sentence):
                a = re.match(regex_list[11], sentence)
                Cung_list.append('')
                Chi_list.append(a[2])
                Sao_list1.append(a[1])
                Sao_list2.append(a[3])
                Sao_list3.append(a[4])
                Can_list.append(a[5])
                continue
            elif re.match(regex_list[12], sentence):
                a = re.match(regex_list[12], sentence)
                Cung_list.append('')
                Chi_list.append(a[2])
                Sao_list1.append(a[1])
                Sao_list2.append('')
                Sao_list3.append(a[3])
                Can_list.append(a[4])
                continue
            elif re.match(regex_list[13], sentence):
                a = re.match(regex_list[13], sentence)
                Cung_list.append('')
                Chi_list.append(a[2])
                Sao_list1.append(a[1])
                Sao_list2.append('')
                Sao_list3.append(a[3])
                Can_list.append(a[4])
                continue
            elif re.match(regex_list[14], sentence):
                a = re.match(regex_list[14], sentence)
                Cung_list.append('')
                Chi_list.append(a[2])
                Sao_list1.append(a[1])
                Sao_list2.append(a[3])
                Sao_list3.append('')
                Can_list.append(a[4])
                continue
            elif re.match(regex_list[15], sentence):
                a = re.match(regex_list[15], sentence)
                Cung_list.append('')
                Chi_list.append(a[2])
                Sao_list1.append(a[1])
                Sao_list2.append(a[3])
                Sao_list3.append('')
                Can_list.append(a[4])
                continue
            elif re.match(regex_list[16], sentence):
                a = re.match(regex_list[16], sentence)
                Cung_list.append('')
                Chi_list.append('')
                Sao_list1.append(a[1])
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append(a[2])
                continue
            elif re.match(regex_list[17], sentence):
                a = re.match(regex_list[17], sentence)
                Cung_list.append('')
                Chi_list.append(a[2])
                Sao_list1.append(a[1])
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[18], sentence):
                a = re.match(regex_list[18], sentence)
                Cung_list.append('')
                Chi_list.append('')
                Sao_list1.append(a[1])
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[19], sentence):
                a = re.match(regex_list[19], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[2])
                Sao_list1.append(a[3])
                Sao_list2.append('')
                Sao_list3.append(a[4])
                Can_list.append('')
                continue
            elif re.match(regex_list[20], sentence):
                a = re.match(regex_list[20], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[2])
                Sao_list1.append(a[3])
                Sao_list2.append(a[4])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[21], sentence):
                a = re.match(regex_list[21], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[2])
                Sao_list1.append(a[3])
                Sao_list2.append(a[4])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[22], sentence):
                a = re.match(regex_list[22], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[2])
                Sao_list1.append(a[3])
                Sao_list2.append(a[4])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[23], sentence):
                a = re.match(regex_list[23], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[2])
                Sao_list1.append(a[3])
                Sao_list2.append(a[4])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[24], sentence):
                a = re.match(regex_list[24], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[2])
                Sao_list1.append(a[3])
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[25], sentence):
                a = re.match(regex_list[25], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[2])
                Sao_list1.append(a[3])
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[26], sentence):
                a = re.match(regex_list[26], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[2])
                Sao_list1.append(a[3])
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[27], sentence):
                a = re.match(regex_list[27], sentence)
                Cung_list.append(a[1])
                Chi_list.append('')
                Sao_list1.append(a[2])
                Sao_list2.append(a[3])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[28], sentence):
                a = re.match(regex_list[28], sentence)
                Cung_list.append(a[1])
                Chi_list.append('')
                Sao_list1.append(a[2])
                Sao_list2.append(a[3])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[29], sentence):
                a = re.match(regex_list[29], sentence)
                Cung_list.append(a[1])
                Chi_list.append('')
                Sao_list1.append(a[2])
                Sao_list2.append(a[3])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[30], sentence):
                a = re.match(regex_list[30], sentence)
                Cung_list.append(a[1])
                Chi_list.append('')
                Sao_list1.append(a[2])
                Sao_list2.append('')
                Sao_list3.append(a[3])
                Can_list.append('')
                continue
            elif re.match(regex_list[31], sentence):
                a = re.match(regex_list[31], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[3])
                Sao_list1.append(a[2])
                Sao_list2.append(a[4])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[32], sentence):
                a = re.match(regex_list[32], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[3])
                Sao_list1.append(a[2])
                Sao_list2.append(a[4])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[33], sentence):
                a = re.match(regex_list[33], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[3])
                Sao_list1.append(a[2])
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[34], sentence):
                a = re.match(regex_list[34], sentence)
                Cung_list.append(a[1])
                Chi_list.append('')
                Sao_list1.append(a[2])
                Sao_list2.append(a[3])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[35], sentence):
                a = re.match(regex_list[35], sentence)
                Cung_list.append(a[1])
                Chi_list.append('')
                Sao_list1.append(a[2])
                Sao_list2.append(a[3])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[36], sentence) and sentence.count("có") != 2:
                a = re.match(regex_list[36], sentence)
                Cung_list.append(a[1])
                Chi_list.append('')
                Sao_list1.append(a[2])
                Sao_list2.append(a[3])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[37], sentence):
                a = re.match(regex_list[37], sentence)
                Cung_list.append(a[1])
                Chi_list.append('')
                Sao_list1.append(a[2])
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[38], sentence):
                a = re.match(regex_list[38], sentence)
                Cung_list.append(a[1])
                Chi_list.append('')
                Sao_list1.append(a[2])
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[39], sentence) and sentence.count("có") != 2:
                a = re.match(regex_list[39], sentence)
                Cung_list.append(a[1])
                Chi_list.append('')
                Sao_list1.append(a[2])
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[40], sentence):
                a = re.match(regex_list[40], sentence)
                Cung_list.append('')
                Chi_list.append(a[2])
                Sao_list1.append(a[1])
                Sao_list2.append(a[3])
                Sao_list3.append(a[4])
                Can_list.append('')
                continue
            elif re.match(regex_list[41], sentence):
                a = re.match(regex_list[41], sentence)
                Cung_list.append('')
                Chi_list.append('')
                Sao_list1.append(a[1])
                Sao_list2.append(a[2])
                Sao_list3.append(a[3])
                Can_list.append('')
                continue
            elif re.match(regex_list[42], sentence):
                a = re.match(regex_list[42], sentence)
                Cung_list.append('')
                Chi_list.append('')
                Sao_list1.append(a[1])
                Sao_list2.append('')
                Sao_list3.append(a[2])
                Can_list.append('')
                continue
            elif re.match(regex_list[43], sentence):
                a = re.match(regex_list[43], sentence)
                Cung_list.append('')
                Chi_list.append(a[2])
                Sao_list1.append(a[1])
                Sao_list2.append(a[3])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[44], sentence):
                a = re.match(regex_list[44], sentence)
                Cung_list.append('')
                Chi_list.append(a[2])
                Sao_list1.append(a[1])
                Sao_list2.append(a[3])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[45], sentence):
                a = re.match(regex_list[45], sentence)
                Cung_list.append('')
                Chi_list.append(a[2])
                Sao_list1.append(a[1])
                Sao_list2.append(a[3])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[46], sentence):
                a = re.match(regex_list[46], sentence)
                Cung_list.append('')
                Chi_list.append(a[2])
                Sao_list1.append(a[1])
                Sao_list2.append(a[3])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[47], sentence):
                a = re.match(regex_list[47], sentence)
                Cung_list.append('')
                Chi_list.append(a[2])
                Sao_list1.append(a[1])
                Sao_list2.append(a[3])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[48], sentence):
                a = re.match(regex_list[48], sentence)
                Cung_list.append('')
                Chi_list.append(a[2])
                Sao_list1.append(a[1])
                Sao_list2.append(a[3])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[49], sentence):
                a = re.match(regex_list[49], sentence)
                Cung_list.append(a[1])
                Chi_list.append(a[2])
                Sao_list1.append('')
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[50], sentence):
                a = re.match(regex_list[50], sentence)
                Cung_list.append('')
                Chi_list.append('')
                Sao_list1.append(a[1])
                Sao_list2.append(a[2])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[51], sentence):
                a = re.match(regex_list[51], sentence)
                Cung_list.append('')
                Chi_list.append('')
                Sao_list1.append(a[1])
                Sao_list2.append(a[2])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[52], sentence) and sentence.count("có") != 2:
                a = re.match(regex_list[52], sentence)
                Cung_list.append('')
                Chi_list.append('')
                Sao_list1.append(a[1])
                Sao_list2.append(a[2])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[53], sentence) and sentence.count("có") != 2:
                a = re.match(regex_list[53], sentence)
                Cung_list.append('')
                Chi_list.append('')
                Sao_list1.append(a[1])
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[54], sentence):
                a = re.match(regex_list[54], sentence)
                Cung_list.append('')
                Chi_list.append('')
                Sao_list1.append(a[1])
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[55], sentence) and sentence.count("và") == 0 and sentence.count("nếu") == 0:
                a = re.match(regex_list[55], sentence)
                Cung_list.append(a[1] + ', ' + a[3])
                Chi_list.append('')
                Sao_list1.append(a[5])
                Sao_list2.append(a[2] + ', ' + a[4])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[56], sentence):
                a = re.match(regex_list[56], sentence)
                Cung_list.append(a[1] + ', ' + a[3])
                Chi_list.append('')
                Sao_list1.append(a[5] + ', ' + a[6])
                Sao_list2.append(a[2] + ', ' + a[4])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[57], sentence):
                a = re.match(regex_list[57], sentence)
                Cung_list.append(a[1] + ', ' + a[3])
                Chi_list.append('')
                Sao_list1.append(a[5])
                Sao_list2.append(a[2] + ', ' + a[4])
                Sao_list3.append('')
                Can_list.append('')
                continue
            elif re.match(regex_list[58], sentence):
                a = re.match(regex_list[58], sentence)
                Cung_list.append(a[1] + ', ' + a[3])
                Chi_list.append('')
                Sao_list1.append(a[5])
                Sao_list2.append(a[2] + ', ' + a[4])
                Sao_list3.append('')
                Can_list.append('')
                continue
            else:
                Cung_list.append('')
                Chi_list.append('')
                Sao_list1.append('')
                Sao_list2.append('')
                Sao_list3.append('')
                Can_list.append('')

        Code =""
        if Cung_list ==['']:
            Code += '0'
        else:
            Code += '1'
        if Chi_list ==['']:
            Code += '0'
        else:
            Code += '1'
        if Can_list ==['']:
            Code += '0'
        else:
            Code += '1'
        if Sao_list1 ==['']:
            Code += '0'
        else:
            Code += '1'
        if Sao_list2 ==['']:
            Code += '0'
        else:
            Code += '1'
        if Sao_list3 ==['']:
            Code += '0'
        else:
            Code += '1'
        return [Code, Cung_list, Chi_list, Can_list, Sao_list1, Sao_list2, Sao_list3]

    """## Giải"""

    def string_process(text):
        list_text = str(text).split(",")
        for i in range(len(list_text)):
            list_text[i] = list_text[i].strip(" ")
            list_text[i] = re.sub(r"(?<=\w)([A-Z,Ă,Â,Đ])", r" \1", list_text[i])
        return list_text

    def GiaiNeu(sentence):
        # sentence = replace_abbreviations(sentence)
        extract_inf = extract(sentence)
        Code = extract_inf[0]
        if Code == '000000':
            pass

        if Code == '000100':
            Sao_list1 = string_process(extract_inf[4][0])
            if "hội họp" in sentence:
                if IsHoiHop(Sao_list1):
                    return True
            if "đồng cung" in sentence:
                if IsDongCung(Sao_list1):
                    return True
            if "vượng địa" in sentence:
                for i in range(len(Sao_list1)):
                    if DacTinh(Sao_list1[i]) != ['Vượng']:
                        return False
                return True
            return False

        if Code == '000101':
            Sao_list1 = string_process(extract_inf[4][0])
            Sao_list3 = string_process(extract_inf[6][0])
            for i in range(len(Sao_list1)):
                for j in range(len(Sao_list3)):
                    temp = []
                    temp.append(str(Sao_list1[i]))
                    temp.append(str(Sao_list3[j]))
                    if IsDongCung(temp):
                        return False
            return True

        if Code == '000110': # Note: còn tuần, triệt
            Sao_list1 = string_process(extract_inf[4][0])
            Sao_list2 = string_process(extract_inf[5][0])
            if "đi với" in sentence.lower():
                a = ['Văn Xương', 'Thiên Sứ']
                b = ['Văn Khúc', 'Thiên Thương']
                if set(a).issubset(Chua("Tật Ách")) and set(b).issubset(Chua("Nô Bộc")):
                    if (('Địa Không'in Chua('Tật Ách') or 'Địa Kiếp'in Chua('Tật Ách')) and ('Kình Dương'in Chua('Tật Ách') or 'Đà La'in Chua('Tật Ách'))):
                        return True
                    elif (('Địa Không'in Chua('Nô Bộc') or 'Địa Kiếp'in Chua('Nô Bộc')) and ('Kình Dương'in Chua('Nô Bộc') or 'Đà La'in Chua('Nô Bộc'))):
                        return True
                return False
            else:
                if "tuần" in sentence.lower():
                    temp = Tuan()
                    if Sao_list1 in temp:
                        return True
                else:
                    Sao_list1.extend(Sao_list2)
                    return IsHoiHop(Sao_list1)
                
                if "triệt" in sentence.lower():
                    temp = Triet()
                    if Sao_list1 in temp:
                        return True
                else:
                    Sao_list1.extend(Sao_list2)
                    return IsHoiHop(Sao_list1)

        if Code == '000111':
            pass

        if Code == '001100':
            Can = string_process(extract_inf[3][0])
            Sao_list1 = string_process(extract_inf[4][0])
            Can_value = CanChi()
            if IsHoiHop(Sao_list1):
                if Can_value[0] in Can:
                    return True
            return False

        if Code == '010100':
            Chi_list = string_process(extract_inf[2][0])
            Sao_list1 = string_process(extract_inf[4][0])
            for i in range(len(Chi_list)):
                DanhSachSao = CungChua(Chi_list[i])
                if set(Sao_list1).issubset(DanhSachSao):
                    return True
            return False

        if Code == '010110':
            Chi_list = string_process(extract_inf[2][0])
            Sao_list1 = string_process(extract_inf[4][0])
            Sao_list2 = string_process(extract_inf[5][0])
            if "triệt" in sentence.lower():
                result = Triet()
                intersection = list(set(Chi_list).intersection(set(result)))
                Saoinchi = []
                for chi in intersection:
                    Saoinchi.append(CungChua(chi))
                combined_list = []
                for sublist in Saoinchi:
                    combined_list.extend(sublist)
                Saoinchi_combined = list(combined_list)
                Sao_list2_combined = list(generate_combinations(Sao_list2))
                if set(Chi_list).intersection(set(result)):
                    for sao in Sao_list2_combined:
                        check = []
                        check.extend(Saoinchi_combined)
                        check.extend(sao)
                        if IsHoiHop(check):
                            return True
            elif "tuần" in sentence.lower():
                result = Tuan()
                intersection = list(set(Chi_list).intersection(set(result)))
                Saoinchi = []
                for chi in intersection:
                    Saoinchi.append(CungChua(chi))
                combined_list = []
                for sublist in Saoinchi:
                    combined_list.extend(sublist)
                Saoinchi_combined = list(combined_list)
                Sao_list2_combined = list(generate_combinations(Sao_list2))
                if set(Chi_list).intersection(set(result)):
                    for sao in Sao_list2_combined:
                        check = []
                        check.extend(Saoinchi_combined)
                        check.extend(sao)
                        if IsHoiHop(check):
                            return True
            else:
                if not IsDongCung(Sao_list1):
                    return False
                elif TaiCung(Sao_list1[0])[0] not in Chi_list:
                    return False

                if "giáp" in sentence.lower():
                    Giap_list = GiapSao(Sao_list1[0])
                    if not set(Sao_list2).issubset(Giap_list):
                        return False
                else:
                    Sao_list1.extend(Sao_list2)
                    if not IsHoiHop(Sao_list1):
                        return False
                return True
            return False
        
        if Code == '010111':
            pass

        if Code == '011100':
            Chi = string_process(extract_inf[2][0])
            Can = string_process(extract_inf[3][0])
            Sao_list1 = string_process(extract_inf[4][0])
            for i in range(len(Sao_list1)):
                if TaiCung(Sao_list1[i])[0] in Chi:
                    temp = CanChi()
                    if temp[0] in Can:
                        return True
            return False

        if Code == '011101':
            pass

        if Code == '011110':
            pass

        if Code == '011111':
            pass

        if Code == '100100': # Note: Cung Thân chưa xử lý
            Cung = string_process(extract_inf[1][0])
            Sao_list1 = string_process(extract_inf[4][0])
            if "tuần" in sentence.lower():
                pass
            elif "triệt" in sentence.lower():
                pass
            elif "chiếu" in sentence.lower():
                if len(Sao_list1) == 2:
                    Sao1 = Sao_list1[0]
                    Sao2 = Sao_list1[1]
                    if Cung[0] not in NamTrongCung(Sao1):
                        return False
                    if Sao2 not in Chieu(Cung[0]):
                        return False
                elif len(Sao_list1) == 1:
                    if Cung[0] not in NamTrongCung(Sao_list1[0]):
                        return False
                return True
            else:
                for i in range(len(Cung)):
                    if 'Thân' == Cung[i]:
                        Cung[i] = CungDongCung()[0]
                        DanhSachSao = Chua(Cung[i])
                    if set(Sao_list1).issubset(DanhSachSao):
                        return True
                return False

        if Code == '100101':
            pass

        if Code == '100110':
            Cung = string_process(extract_inf[1][0])
            Sao_list1 = string_process(extract_inf[4][0])
            Sao_list2 = string_process(extract_inf[5][0])
            DanhSachSao = Chua(Cung[0])
            DanhSachSao_Than = Chua(CungDongCung()[0])
            if "hội họp" in sentence.lower():
                for i in [0, 1]:
                    temp = [str(Sao_list1[i]), str(Sao_list2[0])]
                    temp1 = [str(Sao_list1[i]), str(Sao_list2[1])]
                    if (Sao_list2[0] in DanhSachSao) and (Sao_list2[1] in DanhSachSao_Than):
                        if IsHoiHop(temp) or IsHoiHop(temp1):
                            return True
                for i in [2, 3]:
                    temp = [str(Sao_list1[i]), str(Sao_list2[0])]
                    temp1 = [str(Sao_list1[i]), str(Sao_list2[1])]
                    if (Sao_list2[0] in DanhSachSao) and (Sao_list2[1] in DanhSachSao_Than):
                        if IsHoiHop(temp) or IsHoiHop(temp1):
                            return True
            else:
                for i in range(len(Sao_list1)):
                    temp = [str(Sao_list1[i]), str(Sao_list2[0])]
                    temp1 = [str(Sao_list1[i]), str(Sao_list2[1])]
                    if (Sao_list2[0] in DanhSachSao) and (Sao_list2[1] in DanhSachSao_Than):
                        if IsDongCung(temp) or IsDongCung(temp1):
                            return True
            return False
        if Code == '101100':
            pass

        if Code == '101110':
            pass

        if Code == '110000':
            pass

        if Code == '110100': # Mệnh, Thân chưa xử lý        # Mệnh, Tử Tức
            Cung = string_process(extract_inf[1][0])
            Chi = string_process(extract_inf[2][0])
            Sao_list1 = string_process(extract_inf[4][0])
            if "Mệnh, Thân" in sentence.lower():
                for i in range(len(Cung)):
                    if DongCung(Cung[i])[0] not in Chi:
                        return False
                pass
            else:
                if len(DongCung(Cung[0])) == 0:
                    return False
                elif DongCung(Cung[0])[0] not in Chi:
                    return False

            if "chiếu" in sentence.lower():
                pass
            elif len(Sao_list1)>=4:
                if not IsHoiHop(Sao_list1):
                    return False
            else:
                if not set(Sao_list1).issubset(Chua(Cung[0])):
                    return False

            return True

        if Code == '110101':
            pass

        if Code == '110110':
            Cung = string_process(extract_inf[1][0])
            Chi = string_process(extract_inf[2][0])
            Sao_list1 = string_process(extract_inf[4][0])
            Sao_list2 = string_process(extract_inf[5][0])
            if "đế vượng" in sentence.lower():
                a = ['Tử Vi', 'Thiên Tướng']
                b = ['Thìn', 'Tuất']
                c = ['Đế Vượng', 'Kình Dương']
                d = ['Phá Quân']
                if set(DongCung("Mệnh")).issubset(b):
                    if set(a).issubset(Chua("Mệnh")):
                        if set(d).issubset(Chua(CungDongCung()[0])):
                            e = []
                            e.extend(c)
                            e.extend(d)
                            if IsHoiHop(e):
                                return True
            else:
                for i in range(len(Cung)):
                    if Cung[i] == "Thân":
                        Cung[i] = CungDongCung()[0]
                    if DongCung(Cung[i])[0] in Chi:
                        if set(Sao_list1).issubset(Chua(Cung[i])):
                            temp = Sao_list1.copy()
                            temp.extend(Sao_list2)
                            if IsDongCung(temp):
                                return True
            return False

        if Code == '111100':
            Cung = string_process(extract_inf[1][0])
            Chi = string_process(extract_inf[2][0])
            Can = string_process(extract_inf[3][0])
            Sao_list1 = string_process(extract_inf[4][0])
            for i in range(len(Cung)):
                if DongCung(Cung[i])[0] in Chi:
                    if set(Sao_list1).issubset(Chua(Cung[i])):
                        temp = CanChi()
                        if temp[0] in Can:
                            return True
            return False

        if Code == '111101':
            pass

        if Code == '111110':
            Cung = string_process(extract_inf[1][0])
            Chi = string_process(extract_inf[2][0])
            Can = string_process(extract_inf[3][0])
            Sao_list1 = string_process(extract_inf[4][0])
            Sao_list2 = string_process(extract_inf[5][0])
            for i in range(len(Cung)):
                if DongCung(Cung[i])[0] in Chi:
                    temp = CanChi()
                    if temp[0] in Can:
                        temp1 = Sao_list1.copy()
                        temp1.extend(Sao_list2)
                        if "triệt" in sentence.lower():
                            if set(Sao_list1).issubset(Chua(Cung[i])):
                                temp2 = Triet()
                                if DongCung(Cung[i])[0] in temp2:
                                    return True
                        else:
                            if set(Sao_list1).issubset(Chua(Cung[i])) and IsHoiHop(temp1):
                                return True
            return False
        return False
    
    path_to_file_im = r"/Users/namnguyen/datn/ontology/data/Information_Extraction.xlsx"
    df_im = pd.read_excel(path_to_file_im, header=None)

    loi_giai_neu = []
    loi_giai_thi = []

    for index, row in df_im.iterrows():
        if index >= 0: 
            loi_giai_neu.append(row[4])
            loi_giai_thi.append(row[5])
    
    csv_file = r"/Users/namnguyen/datn/ontology/data/Mapping.csv"
    abbrev_to_full, full_names = load_abbreviations(csv_file)
    
    loigiai = []
    for index, sentence in enumerate(loi_giai_neu):
        sentence_full = replace_abbreviations(sentence, abbrev_to_full, full_names)
        try:
            if GiaiNeu(sentence_full) == True:
                tuple_data = (sentence_full, loi_giai_thi[index])
                loigiai.append(tuple_data)
        except:
            pass

    return loigiai
     

def load_abbreviations(csv_file):
    df = pd.read_csv(csv_file, encoding='utf-8-sig')
    abbrev_to_full = dict(zip(df['Abbreviation'], df['Full Name']))
    full_names = set(df['Full Name'])
    return abbrev_to_full, full_names

def replace_abbreviations(sentence, abbrev_to_full, full_names):
    words = re.findall(r'\b\w+\b|\,', sentence)
    updated_sentence = ""
    i = 0
    while i < len(words):
        if words[i] == ',':
            updated_sentence = updated_sentence.rstrip() + words[i] + " "
            i += 1
            continue

        # Kiểm tra cặp từ
        if i < len(words) - 1 and words[i+1] != ',':
            two_word_combo = words[i] + " " + words[i+1]
            if two_word_combo in full_names:
                updated_sentence += two_word_combo
                if two_word_combo.istitle() and (i+2 < len(words) and words[i+2].istitle()):
                    updated_sentence = updated_sentence.rstrip() + ","
                updated_sentence += " "
                i += 2
                continue

        # Thay thế từ viết tắt và thêm dấu phẩy nếu cần
        if words[i] in abbrev_to_full:
            updated_sentence += abbrev_to_full[words[i]]
            if words[i].istitle() and (i+1 < len(words) and words[i+1].istitle()):
                updated_sentence = updated_sentence.rstrip() + ","
            updated_sentence += " "
        else:
            updated_sentence += words[i]
            if words[i].istitle() and (i+1 < len(words) and words[i+1].istitle()):
                updated_sentence = updated_sentence.rstrip() + ","
            updated_sentence += " "

        i += 1

    return updated_sentence.strip()

def api_giaidoan(request):
    # Database
    conn = pyodbc.connect(
    'Driver={ODBC Driver 18 for SQL Server};'
    'Server=localhost,1433;'  # localhost và cổng 1433
    'Database=TuVi;'
    'UID=sa;'
    'PWD=namthanh99@A;'
    'Encrypt=yes;'
    'TrustServerCertificate=yes;')
    cursor = conn.cursor()
    cursor.execute('update GD_RulesInput set marker = null')
    #cursor.execute('update TuVi_RulesInput set marker = null')
    #cursor.execute('update NuMenh_RulesInput set marker = null')
    cursor.commit()
    #db['diaban'] = list(db['diaban'])
    ## đánh dấu các tập luật thỏa mãn lá số theo thập nhị cung
    # CungSao: Sao, CungChu: cung chủ, CungSo : địa chi(theo thứ tự 1...12: Tý, Sửu...Hợi)
    for index in range(1,len(a)):
        cungSao = a[index].cungSao
        cursor.execute('select Cung_Id from Cung where tenCung = ?',a[index].cungChu)
        cungId = cursor.fetchval()  
        cursor.execute('update GD_RulesInput set marker = 1 where Cung = ? and Chi = ? and Sao is NULL', cungId, a[index].cungSo)
        voChinhDieu = True
        for sao in cungSao:            
            if sao["saoLoai"] == 1:
                voChinhDieu = False
            cursor.execute('update  GD_RulesInput set marker = 1 where Cung = ? and Chi = ? and Sao = ?', cungId, a[index].cungSo,sao['saoID'])
            cursor.execute('update  GD_RulesInput set marker = 1 where Cung is NULL and Chi = ? and Sao = ?', a[index].cungSo,sao['saoID'])
            cursor.execute('update  GD_RulesInput set marker = 1 where Cung = ? and Chi is NULL and Sao = ?', cungId, sao['saoID'])
        if voChinhDieu:         #Đánh dấu khi vô chính diệu (không có sao chính tinh và k xác định được địa chi, có mỗi Cung chủ, trong tập data này k xảy ra trường hợp này)
            cursor.execute('update  GD_RulesInput set marker = 0.3 where Cung = ? and Chi is NULL and Sao is NULL', cungId)
        ## đánh dấu các tập luật của cung thân
        if a[index].cungThan == True:
            cursor.execute("update  GD_RulesInput set marker = 1 where Cung = (select Cung_Id from Cung where tenCung = N'Thân') and Chi = ? and Sao is NULL", a[index].cungSo)
            # đồng cung
            cursor.execute("update  GD_RulesInput set marker = 1 where Cung = (select Cung_Id from Cung where tenCung = N'Thân') and Chi is NULL and Sao is NULL")
            cursor.execute('update  GD_RulesInput set marker = 1 where Cung = ? and Chi is NULL and Sao is NULL', cungId)

            for sao in cungSao:
                cursor.execute("update  GD_RulesInput set marker = 1 where Cung = (select Cung_Id from Cung where tenCung = N'Thân') and Chi = ? and Sao = ?", a[index].cungSo,sao['saoID'])
                cursor.execute('update  GD_RulesInput set marker = 1 where Cung is NULL and Chi = ? and Sao = ?', a[index].cungSo,sao['saoID'])
                cursor.execute("update  GD_RulesInput set marker = 1 where Cung = (select Cung_Id from Cung where tenCung = N'Thân') and Chi is NULL and Sao = ?", sao['saoID'])
            ## đánh dấu tập luật cho triệt và tuần của cung thân nếu có
            if a[index].anTriet ==True:
                cursor.execute("update  GD_RulesInput set marker = 1 where Cung = ? and Chi = ? and Sao = (select Sao_Id from Sao where saoTen = N'Triệt')", cungId, a[index].cungSo)
                cursor.execute("update  GD_RulesInput set marker = 1 where Cung = (select Cung_Id from Cung where tenCung = N'Thân') and Chi is NULL and Sao = (select Sao_Id from Sao where saoTen = N'Tuần')")
            elif a[index].anTuan == True:
                cursor.execute("update  GD_RulesInput set marker = 1 where Cung = ? and Chi = ? and Sao = (select Sao_Id from Sao where saoTen = N'Tuần')", cungId, a[index].cungSo)
                cursor.execute("update  GD_RulesInput set marker = 1 where Cung = (select Cung_Id from Cung where tenCung = N'Thân') and Chi is NULL and Sao = (select Sao_Id from Sao where saoTen = N'Tuần')")
                
        ## đánh dấu tập luật cho triệt và tuần
        if a[index].anTriet ==True:
            cursor.execute("update  GD_RulesInput set marker = 1 where Cung = ? and Chi = ? and Sao = (select Sao_Id from Sao where saoTen = N'Triệt')", cungId, a[index].cungSo)
            cursor.execute("update  GD_RulesInput set marker = 1 where Cung = ? and Chi is NULL and Sao = (select Sao_Id from Sao where saoTen = N'Triệt')", cungId)
        elif a[index].anTuan == True:
            cursor.execute("update  GD_RulesInput set marker = 1 where Cung = ? and Chi = ? and Sao = (select Sao_Id from Sao where saoTen = N'Tuần')", cungId, a[index].cungSo)
            cursor.execute("update  GD_RulesInput set marker = 1 where Cung = ? and Chi is NULL and Sao = (select Sao_Id from Sao where saoTen = N'Tuần')", cungId)
        cursor.commit()

    cursor1 = conn.cursor()
    cursor1.execute('update TuVi_RulesInput set marker = null')
    cursor1.commit()
    menhCung = ''
    for index in range(1,len(a)):
        if a[index].cungChu == 'Mệnh':
            menhCung = a[index].cungTen

    for index in range(1,len(a)):
        cungSao1 = a[index].cungSao
        for sao in cungSao1:
            if sao['saoID'] == 1:
                cursor1.execute("update TuVi_RulesInput set marker = 1 where cung = (select Chi_Id from Chi where tenChi = ?) and cung1 = (select Chi_Id from Chi where tenChi = ?)", a[index].cungTen, menhCung)
        cursor1.commit()

    cursor2 = conn.cursor()
    cursor2.execute('update NuMenh_RulesInput set marker = null')
    cursor2.commit()
    if b.gioiTinh == -1:
        for index in range(1,len(a)):
            if a[index].cungChu == 'Mệnh':
                cungSao2 = a[index].cungSao
                for sao in cungSao2:
                    cursor2.execute("update NuMenh_RulesInput set marker = 1 where Sao = ?", sao['saoID'])
                cursor2.commit()

    cursor3 = conn.cursor()
    cursor3.execute('update C6_RulesInput set marker = null')
    cursor3.commit()
    for index in range(1,len(a)):
        cungSao3 = a[index].cungSao
        cursor3.execute('select Cung_Id from Cung where tenCung = ?',a[index].cungChu)
        cungId1 = cursor3.fetchval() 
        voChinhDieu1 = True
        for sao in cungSao3:
            if sao["saoLoai"] == 1:
                voChinhDieu1 = False
            cursor3.execute('update C6_RulesInput set marker = 2 where Cung = ? and Sao = ?', cungId1, sao['saoID'])
        if voChinhDieu1:         
            cursor3.execute('update C6_RulesInput set marker = 0.3 where Cung = ? and Sao is NULL', cungId)
        cursor3.commit()

    cursor4 = conn.cursor()
    cursor4.execute('update C6_Tho_Rules set marker = 3')
    cursor4.commit()

    cursor5 = conn.cursor()
    cursor5.execute('update NgheNghiep_RulesInput set marker = null')
    cursor5.execute('update NgheNghiep_RulesInput set marker = 4 where Cuc = (select Hanh_Id from Hanh where tenHanh = ?) and NgayAL = ? and cungMTB = (select Chi_Id from Chi where tenChi = ?)', b.tenHanh, b.ngayAm, menhCung)
    cursor5.commit()

    giai_doan = cursor.execute('exec GetDataGD').fetchall()
    giai_doan1 = cursor1.execute('exec GetDataTuVi').fetchall()
    giai_doan2 = cursor2.execute('exec GetDataNuMenh').fetchall()
    giai_doan3 = cursor3.execute('exec GetDataC6V1').fetchall()
    giai_doan4 = cursor4.execute('exec GetDataC6Tho').fetchall()
    giai_doan5 = cursor5.execute('exec GetDataNgheNghiep').fetchall()
    # print(giai_doan[0])
    # print(giai_doan1[0])
    # print(giai_doan2[0])
    # print(giai_doan3[0])
    # print(giai_doan4)
    # print(giai_doan5)
    # giaidoan = []
    # for item in giai_doan:
    #     giaidoan.append(list(item))

    # for item1 in giai_doan1:
    #     giaidoan.append(list(item1))

    # for item2 in giai_doan2:
    #     giaidoan.append(list(item2))
    
    # for item3 in giai_doan3:
    #     giaidoan.append(list(item3))
    
    # for item4 in giai_doan4:
    #     giaidoan.append(list(item4))

    # for item5 in giai_doan5:
    #     giaidoan.append(list(item5))
    giaidoan = []
    def append_if_not_exists(source, dest):
        for item in source:
            if list(item) not in dest:
                dest.append(list(item))

    # Sử dụng hàm vừa định nghĩa cho mỗi giai đoạn
    append_if_not_exists(giai_doan, giaidoan)
    append_if_not_exists(giai_doan1, giaidoan)
    append_if_not_exists(giai_doan2, giaidoan)
    append_if_not_exists(giai_doan3, giaidoan)
    append_if_not_exists(giai_doan4, giaidoan)
    append_if_not_exists(giai_doan5, giaidoan)
    
    # print(giaidoan)
    giaidoann = []
    if 'Tất cả' in tk:
        giaidoann = giaidoan
    else:
        for i in tk:
            for rl in giaidoan:
                if str(i) in str(rl):
                    giaidoann.append(rl)
    # print(giaidoann)
    giaidoannn = unique_list(giaidoann)
    # giaidoannn = json.dumps(giaidoannn)
    
    # Ontology
    path_to_file1 = r"/Users/namnguyen/datn/ontology/data/diaban.csv"
    df_dia_ban = pd.read_csv(path_to_file1, header=None)
    path_to_file2 = r"/Users/namnguyen/datn/ontology/data/thienban.csv"
    df_thienBan = pd.read_csv(path_to_file2, header=None)
    path_to_file_im = r"/Users/namnguyen/datn/ontology/data/Information_Extraction.xlsx"
    df_im = pd.read_excel(path_to_file_im, header=None)

    kq_giai = process_ontology_data(df_dia_ban, df_thienBan, df_im)
    # print(kq_giai)
    giaidoan = [list(item) for item in kq_giai]
    tuples_to_check = ['mệnh', 'thân', 'quan lộc', 'tài bạch', 'thiên di', 'phúc đức', 'phu thê', 
                   'tử tức', 'điền trạch', 'tật ách', 'phụ mẫu', 'huynh đệ', 'nô bộc']
    modified_giaidoan = []
    for tup in giaidoan:
        first_part = tup[0].lower()
        keyword_found = None
        for keyword in tuples_to_check:
            if keyword in first_part:
                keyword_found = "CUNG " + keyword.upper()
        if not keyword_found:
            keyword_found = "GIẢI ĐOÁN CHUNG"
        new_tuple = (keyword_found,) + tuple(tup) + ("Tử Vi Áo Bí",)
        modified_giaidoan.append(new_tuple)
    giaidoan_json = modified_giaidoan + giaidoannn
    giaidoan_json = json.dumps(giaidoan_json)
       
    my_return = (json.dumps(giaidoan_json, default=lambda o: o.__dict__))
    return HttpResponse(my_return, content_type="application/json")
    
def api_daivan(request):
    conn = pyodbc.connect(
    'Driver={ODBC Driver 18 for SQL Server};'
    'Server=localhost,1433;'  # localhost và cổng 1433
    'Database=TuVi;'
    'UID=sa;'
    'PWD=namthanh99@A;'
    'Encrypt=yes;'
    'TrustServerCertificate=yes;')

    cursor = conn.cursor()
    cursor.execute('update DV_RulesInput set marker = null')
    cursor.execute("TRUNCATE TABLE RuleforDV")
    cursor.commit()

    for index in range(1,len(a)):
        cursor.execute("update DV_RulesInput set marker = ? where Ban_menh = ? and Hanh_dai_van = (select Hanh_Id from Hanh where tenHanh = ?)", index, b.hanhCuc, a[index].hanhCung)
        cursor.execute("update DV_RulesInput set marker = ? where Tam_hop_tuoi = ? and Tam_hop_dai_van = (select Chi_Id from Chi where tenChi = ?)", index, b.chiNam, a[index].cungTieuHan)
        cursor.execute("INSERT INTO RuleforDV select mucluc, nguyenhan, Mucdo, Tham_khao, marker, DV_RuleInsId from DV_RulesInput, DV_Rules where DV_RuleId = DV_RuleInsId and marker = ?", index)
        cursor.commit()

    for index in range(1,len(a)):
        #cursor.execute('update DV_Rule set mucluc = ? where rule_Id in (select ruleID from DV_RuleInputs where marker = ?)', str('Đại vận ' + str(a[index].cungDaiHan) + '-' + str(a[index].cungDaiHan+9)), index)
        cursor.execute('update RuleforDV set mucluc = ? where marker = ?', str('ĐẠI VẬN ' + str(a[index].cungDaiHan) + '-' + str(a[index].cungDaiHan+9)), index)
    cursor.commit()
    
    giai_doan1 = cursor.execute('select * from RuleforDV').fetchall()
    # print(giai_doan1[0])
    giaidoan1 = []
    for item in giai_doan1:
        giaidoan1.append(list(item))
    # print(giaidoan1)
    giaidoann1 = []

    if tk == "Tất cả":
        giaidoann1 = giaidoan1
    else:
        giaidoann1 = []
        for rl in giaidoan1:
            if tk in str(rl): # Bạn cần kiểm tra điều kiện này dựa trên cách bạn muốn so sánh
                giaidoann1.append(rl)
        # for i in tk:
        #     for rl in giaidoan1:
        #         if str(i) in str(rl):
        #             giaidoann1.append(rl)

    giaidoannn1 = unique_list(giaidoann1)
    giaidoannn1 = json.dumps(giaidoannn1)
    # print(giaidoannn1)
    my_return = (json.dumps(giaidoannn1, default=lambda o: o.__dict__))
    return HttpResponse(my_return, content_type="application/json")

def api_vanhan(request):
    conn = pyodbc.connect(
    'Driver={ODBC Driver 18 for SQL Server};'
    'Server=localhost,1433;'  # localhost và cổng 1433
    'Database=TuVi;'
    'UID=sa;'
    'PWD=namthanh99@A;'
    'Encrypt=yes;'
    'TrustServerCertificate=yes;')
    cursor = conn.cursor()
    cursor.execute('update VH_RulesInput set marker = null')
    cursor.commit()

    for index in range(1,len(a)):
        if a[index].cungTieuHan == b.chiNamXHTen:
            cursor.execute("update VH_RulesInput set marker = 1 where Ban_menh = ? and Cung_tieu_han = (select Chi_Id from Chi where tenChi = ?)", b.hanhCuc, a[index].cungTieuHan)
            for sao in a[index].cungSao:
                cursor.execute('update  VH_RulesInput set marker = 1 where Cung_tieu_han = (select Chi_Id from Chi where tenChi = ?) and Sao1 = ? and Sao2 = ?', a[index].cungTieuHan, sao['saoID'], sao['saoID'])
                cursor.execute('update  VH_RulesInput set marker = 1 where Cung_tieu_han = (select Chi_Id from Chi where tenChi = ?) and Sao1 = ? and Sao2 is NULL', a[index].cungTieuHan, sao['saoID'])
                cursor.execute('update  VH_RulesInput set marker = 1 where Cung_tieu_han is NULL and Sao1 = ? and Sao2 = ?', sao['saoID'], sao['saoID'])
                cursor.execute('update  VH_RulesInput set marker = 1 where Cung_tieu_han is NULL and Sao1 = ? and Sao2 is NULL', sao['saoID'])
        cursor.commit()

    giai_doan2 = cursor.execute('exec GetDataVH').fetchall()
    # print(giai_doan2)
    giaidoan2 = []
    for item in giai_doan2:
        giaidoan2.append(list(item))

    giaidoann2 = []
    if tk == "Tất cả":
        giaidoann2 = giaidoan2
    else:
        giaidoann2 = []
        for rl in giaidoan2:
            if tk in str(rl): # Bạn cần kiểm tra điều kiện này dựa trên cách bạn muốn so sánh
                giaidoann2.append(rl)
    
    giaidoannn2 = unique_list(giaidoann2)
    giaidoannn2 = json.dumps(giaidoannn2)
    # print(giaidoannn2)
    my_return = (json.dumps(giaidoannn2, default=lambda o: o.__dict__))
    return HttpResponse(my_return, content_type="application/json")

def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)  
    copy_sheet_attributes(source_sheet, target_sheet)


def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])
    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) 
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)


def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)
        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)
        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)


def api_trondoi(request):
    wb_result = openpyxl.Workbook()
    wb = load_workbook("C:/Users/vanhu/Desktop/ONTO_LSTV/NguyenNgocThin20185408/Code/lasotuvi_django/template.xlsx", data_only=True)

    ws3 = wb_result.create_sheet('Thông tin')
    copy_sheet(wb['Thông tin'], ws3)
    ws3[get_column_letter(5) + str(4)] = b.ten
    ws3[get_column_letter(5) + str(5)] = b.namNu
    ws3[get_column_letter(3) + str(6)] = b.ngayDuong
    ws3[get_column_letter(4) + str(6)] = b.thangDuong
    ws3[get_column_letter(5) + str(6)] = b.namDuong
    ws3[get_column_letter(5) + str(7)] = b.gioSinh
    ws3[get_column_letter(5) + str(8)] = b.namXemHan
    ws3[get_column_letter(9) + str(7)] = b.timeZone

    ws = wb_result.create_sheet('Tóm lược trọn đời')
    copy_sheet(wb['Tóm lược trọn đời'], ws)

    for row in range(4,54):
        ws[get_column_letter(2) + str(row)] = int(ws[get_column_letter(3) + str(row)].value - b.namAm)
    
    for index in range(1, len(a)):
        for row in range(4,54):
            if ws[get_column_letter(5) + str(row)].value == a[index].cungTieuHan:
                ws[get_column_letter(6) + str(row)] = str(a[index].cungChu)
                ws[get_column_letter(100) + str(row)] = str(a[index].hanhCung)
            if ws[get_column_letter(2) + str(row)].value in range(int(a[index].cungDaiHan), int(a[index].cungDaiHan+10)):
                ws[get_column_letter(1) + str(row)] = 'ĐẠI VẬN ' + str(a[index].cungDaiHan) + '-' + str(a[index].cungDaiHan+9)
    
    for row in range(4,54):
        if ws[get_column_letter(6) + str(row)].value == 'Điền trạch':
            ws[get_column_letter(7) + str(row)] = 'Tình hình đất đai, nhà cửa'
        elif ws[get_column_letter(6) + str(row)].value == 'Quan lộc':
            ws[get_column_letter(7) + str(row)] = 'Công danh sự nghiệp'
        elif ws[get_column_letter(6) + str(row)].value == 'Nô bộc':
            ws[get_column_letter(7) + str(row)] = 'Liên quan đến bạn bè, đồng nghiệp'
        elif ws[get_column_letter(6) + str(row)].value == 'Thiên di':
            ws[get_column_letter(7) + str(row)] = 'Tình trạng ngoại cảnh, môi trường sống'
        elif ws[get_column_letter(6) + str(row)].value == 'Tật Ách':
            ws[get_column_letter(7) + str(row)] = 'Tình trạng sức khỏe'
        elif ws[get_column_letter(6) + str(row)].value == 'Tài Bạch':
            ws[get_column_letter(7) + str(row)] = 'Tài sản, tiền bạc'
        elif ws[get_column_letter(6) + str(row)].value == 'Tử tức':
            ws[get_column_letter(7) + str(row)] = 'Tình trạng con, cháu'
        elif ws[get_column_letter(6) + str(row)].value == 'Phu thê':
            ws[get_column_letter(7) + str(row)] = 'Tình trạng hôn nhân, đời sống tình cảm'
        elif ws[get_column_letter(6) + str(row)].value == 'Huynh đệ':
            ws[get_column_letter(7) + str(row)] = 'Liên quan đến anh, chị, em trong gia đình'
        elif ws[get_column_letter(6) + str(row)].value == 'Phụ mẫu':
            ws[get_column_letter(7) + str(row)] = 'Liên quan đến cha, mẹ'
        elif ws[get_column_letter(6) + str(row)].value == 'Mệnh':
            ws[get_column_letter(7) + str(row)] = 'Bản mệnh của mình'
        elif ws[get_column_letter(6) + str(row)].value == 'Phúc đức':
            ws[get_column_letter(7) + str(row)] = 'Tư tưởng tinh thần, phúc đức'
        else:
            continue

    red = Color(rgb='FF0000')
    yellow = Color(rgb='FFFF00')
    green = Color(rgb='00FF00')
    fill_red = PatternFill(patternType='solid', fgColor=red)
    fill_yellow = PatternFill(patternType='solid', fgColor=yellow)
    fill_green = PatternFill(patternType='solid', fgColor=green)

    for row in range(4,54):
        if (b.tenHanh == 'Hỏa') and (ws[get_column_letter(100) + str(row)].value == 'Kim'):
            ws[get_column_letter(8) + str(row)] = 'Đỏ'
            ws[get_column_letter(8) + str(row)].fill = fill_red
            ws[get_column_letter(9) + str(row)] = 'Khó mà tránh thoát được những tai ương khủng khiếp'
        elif (b.tenHanh == 'Kim') and (ws[get_column_letter(100) + str(row)].value == 'Thủy'):
            ws[get_column_letter(8) + str(row)] = 'Đỏ'
            ws[get_column_letter(8) + str(row)].fill = fill_red
            ws[get_column_letter(9) + str(row)] = 'Thường mắc tai ương, rất đáng lo ngại'
        elif (b.tenHanh == 'Mộc') and (ws[get_column_letter(100) + str(row)].value == 'Hỏa'):
            ws[get_column_letter(8) + str(row)] = 'Đỏ'
            ws[get_column_letter(8) + str(row)].fill = fill_red
            ws[get_column_letter(9) + str(row)] = 'Rất khó tránh những tai ương họa hại'
        elif (b.tenHanh == 'Thủy') and (ws[get_column_letter(100) + str(row)].value == 'Mộc'):
            ws[get_column_letter(8) + str(row)] = 'Đỏ'
            ws[get_column_letter(8) + str(row)].fill = fill_red
            ws[get_column_letter(9) + str(row)] = 'Mọi việc đều bị trắc trở, không được xứng ý toại lòng'
        elif (b.tenHanh == 'Thổ') and (ws[get_column_letter(100) + str(row)].value == 'Mộc'):
            ws[get_column_letter(8) + str(row)] = 'Đỏ'
            ws[get_column_letter(8) + str(row)].fill = fill_red
            ws[get_column_letter(9) + str(row)] = 'Thường hay mắc bệnh điên cuồng hay bệnh khí huyết rất nguy hiểm'
        elif ((b.tenHanh == 'Kim') and (ws[get_column_letter(100) + str(row)].value == 'Thổ')) or ((b.tenHanh == "Mộc") and (ws[get_column_letter(100) + str(row)].value == 'Thủy')) or ((b.tenHanh == "Thổ") and (ws[get_column_letter(100) + str(row)].value == 'Kim') or ((b.tenHanh == "Hỏa") and (ws[get_column_letter(100) + str(row)].value == 'Mộc')) or ((b.tenHanh == "Thổ") and (ws[get_column_letter(100) + str(row)].value == 'Hỏa'))):
            ws[get_column_letter(8) + str(row)] = 'Xanh'
            ws[get_column_letter(8) + str(row)].fill = fill_green
            ws[get_column_letter(9) + str(row)] = 'Tốt'
        else:
            ws[get_column_letter(8) + str(row)] = 'Vàng'
            ws[get_column_letter(8) + str(row)].fill = fill_yellow
            ws[get_column_letter(9) + str(row)] = 'Bình thường'

    ws1 = wb_result.create_sheet('Theo 10 năm')
    copy_sheet(wb['Theo 10 năm'], ws1)

    conn = pyodbc.connect(
    'Driver={ODBC Driver 18 for SQL Server};'
    'Server=localhost,1433;'  # localhost và cổng 1433
    'Database=TuVi;'
    'UID=sa;'
    'PWD=namthanh99@A;'
    'Encrypt=yes;'
    'TrustServerCertificate=yes;')
    cursor = conn.cursor()
    row_lg1 = cursor.execute('select * from RuleforDV').fetchall()
    list_DV = []
    columnNames1 = [column[0] for column in cursor.description]
    for record in row_lg1:
        list_DV.append(dict(zip(columnNames1,record)))

    for row in range(4,54):
        if b.namXemHan == int(ws[get_column_letter(3) + str(row)].value):
            ws1[get_column_letter(1) + str(3)] = ws[get_column_letter(1) + str(row)].value

    k1 = 5
    Dv = []
    for i in range(len(list_DV)):
        if list_DV[i]['mucluc'] == ws1[get_column_letter(1) + str(3)].value:
            Dv.append(list_DV[i])

    for i in range(len(Dv)):
        # ws1[get_column_letter(2) + str(k1+i)] = str(list_DV[i]['nguyenhan']) + ' thì ' + str(list_DV[i]['mucdo'])
        ws1[get_column_letter(2) + str(k1+i)] = str(list_DV[i].get('nguyenhan', 'Không có nguyên hạn')) + ' thì ' + str(list_DV[i].get('mucdo', 'Không có mức độ'))

    for index in range(1,len(a)):
        for row in range(4,54):
            if b.namXemHan == int(ws[get_column_letter(3) + str(row)].value) and (ws[get_column_letter(2) + str(row)].value in range(int(a[index].cungDaiHan), int(a[index].cungDaiHan+10))):
                ws1[get_column_letter(5) + str(3)] = 'Năm ' + str(b.namDuong + a[index].cungDaiHan) + '-' + str(b.namDuong + a[index].cungDaiHan + 9)
            if (b.namXemHan - b.namAm) in range(int(a[index].cungDaiHan), int(a[index].cungDaiHan+10)):
                ws1[get_column_letter(100) + str(3)] = a[index].hanhCung

    if ((b.tenHanh == 'Hỏa') and (ws1[get_column_letter(100) + str(3)].value == 'Kim')) or ((b.tenHanh == 'Kim') and (ws1[get_column_letter(100) + str(3)].value == 'Thủy')) or ((b.tenHanh == 'Mộc') and (ws1[get_column_letter(100) + str(3)].value == 'Hỏa')) or ((b.tenHanh == 'Thủy') and (ws1[get_column_letter(100) + str(3)].value == 'Mộc')) or ((b.tenHanh == 'Thổ') and (ws1[get_column_letter(100) + str(3)].value == 'Mộc')):
        ws1[get_column_letter(9) + str(3)] = 'Đỏ'
        ws1[get_column_letter(9) + str(3)].fill = fill_red
    elif ((b.tenHanh == 'Kim') and (ws1[get_column_letter(100) + str(3)].value == 'Thổ')) or ((b.tenHanh == "Mộc") and (ws1[get_column_letter(100) + str(3)].value == 'Thủy')) or ((b.tenHanh == "Thổ") and (ws1[get_column_letter(100) + str(3)].value == 'Kim') or ((b.tenHanh == "Hỏa") and (ws1[get_column_letter(100) + str(3)].value == 'Mộc')) or ((b.tenHanh == "Thổ") and (ws1[get_column_letter(100) + str(3)].value == 'Hỏa'))):
        ws1[get_column_letter(9) + str(3)] = 'Xanh'
        ws1[get_column_letter(9) + str(3)].fill = fill_green
    else:
        ws1[get_column_letter(9) + str(3)] = 'Vàng'
        ws1[get_column_letter(9) + str(3)].fill = fill_yellow
    
    list_row = []
    for row in range(4,54):
        if ws1[get_column_letter(1) + str(3)].value == ws[get_column_letter(1) + str(row)].value:
            list_row.append(row)
    
    k = 11
    for i in range(len(list_row)):
        ws1[get_column_letter(1) + str(k+i)] = ws[get_column_letter(2) + str(list_row[i])].value
        ws1[get_column_letter(2) + str(k+i)] = ws[get_column_letter(3) + str(list_row[i])].value
        ws1[get_column_letter(3) + str(k+i)] = ws[get_column_letter(4) + str(list_row[i])].value
        ws1[get_column_letter(4) + str(k+i)] = ws[get_column_letter(5) + str(list_row[i])].value
        ws1[get_column_letter(5) + str(k+i)] = ws[get_column_letter(6) + str(list_row[i])].value
        ws1[get_column_letter(6) + str(k+i)] = ws[get_column_letter(7) + str(list_row[i])].value
        ws1[get_column_letter(7) + str(k+i)] = ws[get_column_letter(8) + str(list_row[i])].value
        if (ws1[get_column_letter(7) + str(k+i)].value == 'Đỏ'):
            ws1[get_column_letter(7) + str(k+i)].fill = fill_red
        elif (ws1[get_column_letter(7) + str(k+i)].value == 'Xanh'):
            ws1[get_column_letter(7) + str(k+i)].fill = fill_green
        else:
            ws1[get_column_letter(7) + str(k+i)].fill = fill_yellow
        ws1[get_column_letter(8) + str(k+i)] = ws[get_column_letter(9) + str(list_row[i])].value
        ws1[get_column_letter(9) + str(k+i)] = ws[get_column_letter(10) + str(list_row[i])].value

    ws2 = wb_result.create_sheet('Theo từng năm')
    copy_sheet(wb['Theo từng năm'], ws2)
    cursor1 = conn.cursor()
    row_lg2 = cursor1.execute('exec GetDataVH').fetchall()
    list_VH = []
    columnNames2 = [column[0] for column in cursor1.description]
    for record in row_lg2:
        list_VH.append(dict(zip(columnNames2,record)))
    k2 = 6
    for i in range(len(list_VH)):
        ws2[get_column_letter(2) + str(k2+i)] = str(list_VH[i]['nguyenhan']) + ' thì ' + str(list_VH[i]['Loi_giai'])

    for row in range(4,54):
        if b.namXemHan == int(ws[get_column_letter(3) + str(row)].value):
            ws2[get_column_letter(1) + str(4)] = 'Năm ' + str(ws[get_column_letter(4) + str(row)].value)
            ws2[get_column_letter(4) + str(4)] = 'Năm ' + str(ws[get_column_letter(3) + str(row)].value)
            ws2[get_column_letter(7) + str(4)] = ws[get_column_letter(8) + str(row)].value
            if (ws2[get_column_letter(7) + str(4)].value == 'Đỏ'):
                ws2[get_column_letter(7) + str(4)].fill = fill_red
            elif (ws2[get_column_letter(7) + str(4)].value == 'Xanh'):
                ws2[get_column_letter(7) + str(4)].fill = fill_green
            else:
                ws2[get_column_letter(7) + str(4)].fill = fill_yellow
    
    key_column = 1
    merge_columns = [1]
    start_row = 4
    max_row = 54
    key = None
    for row, row_cells in enumerate(ws.iter_rows(min_col=key_column, min_row=start_row,max_col=key_column, max_row=max_row),start_row):
        if key != row_cells[0].value or row == max_row:
            if not key is None:
                for merge_column in merge_columns:
                    ws.merge_cells( start_row=start_row, start_column=merge_column,end_row=row - 1, end_column=merge_column)
                    ws.cell(row=start_row, column=merge_column).alignment = Alignment(horizontal='center', vertical='center')
                start_row = row
            key = row_cells[0].value
        if row == max_row: row += 1 

    if 'Sheet' in wb_result.sheetnames:  
        wb_result.remove(wb_result['Sheet'])

    response = HttpResponse(content_type='application/vnd.ms-excel')
    path = str(str(b.ten) + '_' + str(b.namDuong) +  str(b.ngayDuong) + str(b.thangDuong) + '.xlsx')
    response['Content-Disposition'] = 'attachment; filename=' + path
    wb_result.save(response)
    #wb_result.save("xlsx-to-pdf.pdf", SaveFormat.PDF)

    # jpype.startJVM() 
    # response1 = HttpResponse(content_type='application/pdf')
    # # response1['Content-Disposition'] = 'inline;filename=result.pdf'
    # # workbook = Workbook(wb_result)
    # # workbook.Save(response1)
    # wb_result.save(response1, SaveFormat.PDF)
    #jpype.shutdownJVM()

    return response

def api_convertPdf(request):
    path = f"C:/Users/vanhu/Desktop/ONTO_LSTV/NguyenNgocThin20185408/Code/lasotuvi_django/lasotuvi1/templates/pdf.html"
    url = 'http://127.0.0.1:8000'

    response = urllib.request.urlopen(url)
    webContent = response.read().decode('UTF-8')

    #f = open(path, 'w')
    #f.write(webContent)
    with open(path, 'w', encoding='utf-8') as f:
        f.write(webContent)
    f.close


    